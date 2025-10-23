'''
Business: Processes Excel files by applying formulas or operations based on user query
Args: event - dict with httpMethod, body (JSON with base64 file, query, language)
      context - object with request_id, function_name attributes
Returns: HTTP response with processed Excel file as base64 or error
'''

import json
import base64
from typing import Dict, Any
import openpyxl
from openpyxl import load_workbook
from io import BytesIO
import os
import anthropic

def apply_excel_modifications(code: str, wb, ws):
    lines = code.strip().split('\n')
    for line in lines:
        if '=' in line and not line.strip().startswith('#'):
            try:
                if 'ws[' in line:
                    parts = line.split('=', 1)
                    target = parts[0].strip()
                    value = parts[1].strip().strip('"').strip("'")
                    
                    if target.startswith('ws['):
                        cell_ref = target[3:-1].strip('"').strip("'")
                        
                        if value.startswith('='):
                            ws[cell_ref] = value
                        elif value.startswith('f'):
                            formula_str = value[2:-1] if value.startswith("f'") or value.startswith('f"') else value[1:]
                            ws[cell_ref] = formula_str
                        else:
                            try:
                                ws[cell_ref] = float(value) if '.' in value else int(value)
                            except:
                                ws[cell_ref] = value
            except:
                pass

def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    method: str = event.get('httpMethod', 'POST')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Max-Age': '86400'
            },
            'body': ''
        }
    
    if method != 'POST':
        return {
            'statusCode': 405,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': 'Method not allowed'})
        }
    
    body_data = json.loads(event.get('body', '{}'))
    file_base64 = body_data.get('file')
    query = body_data.get('query', '')
    language = body_data.get('language', 'ru')
    
    if not file_base64 or not query:
        return {
            'statusCode': 400,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': 'Missing file or query'})
        }
    
    file_bytes = base64.b64decode(file_base64)
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active
    
    data_preview = []
    for idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        if idx < 5:
            data_preview.append([str(cell) if cell is not None else '' for cell in row])
    
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return {
            'statusCode': 500,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': 'API key not configured'})
        }
    
    client = anthropic.Anthropic(api_key=api_key)
    
    system_prompt = """You are an Excel processing assistant. Given a user query and Excel data preview, 
provide Python code using openpyxl to modify the workbook. The code should:
1. Use variable 'wb' (already loaded workbook) and 'ws' (active sheet)
2. Modify cells, add formulas, or perform calculations as requested
3. Be safe and not delete all data unless explicitly asked
4. Return only executable Python code, no explanations

Example response:
```python
ws['D1'] = 'Total'
for row in range(2, ws.max_row + 1):
    ws[f'D{row}'] = f'=B{row}*C{row}'
```"""

    user_prompt = f"User query: {query}\n\nData preview (first 5 rows):\n{json.dumps(data_preview, ensure_ascii=False)}\n\nProvide Python code to process this Excel file."
    
    message = client.messages.create(
        model="claude-3-5-sonnet-20241022",
        max_tokens=1024,
        messages=[{"role": "user", "content": user_prompt}],
        system=system_prompt
    )
    
    code = message.content[0].text
    code = code.replace('```python', '').replace('```', '').strip()
    
    apply_excel_modifications(code, wb, ws)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    result_base64 = base64.b64encode(output.read()).decode('utf-8')
    
    return {
        'statusCode': 200,
        'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
        'isBase64Encoded': False,
        'body': json.dumps({
            'file': result_base64,
            'message': 'File processed successfully'
        })
    }